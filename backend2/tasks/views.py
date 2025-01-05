# tasks/views.py
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from django.conf import settings
from .models import Task
from .serializers import TaskSerializer
import logging
from users.authenticate import NodeJWTAuthentication
from rest_framework_simplejwt.exceptions import AuthenticationFailed
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

logger = logging.getLogger(__name__)

class TaskListCreateView(APIView):
    authentication_classes = [NodeJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tasks = Task.objects.filter(user=request.user)
        serializer = TaskSerializer(tasks, many=True, context={'request': request})
        return Response(serializer.data)

    def post(self, request):
        logger.debug(f"Authenticated user: {request.user}")  # Log the authenticated user

        serializer = TaskSerializer(data=request.data, context={'request': request})
        
        if serializer.is_valid():
            serializer.save()  # The user should be automatically set now
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        logger.debug(f"Serializer errors: {serializer.errors}")  # Log the errors if serializer is not valid
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    
class TaskDetailView(APIView):
    authentication_classes = [NodeJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            task = Task.objects.get(id=pk, user=request.user)
        except Task.DoesNotExist:
            logger.error(f"Task with id {pk} not found for user {request.user}")
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = TaskSerializer(task)
        return Response(serializer.data)

    def put(self, request, pk):
        try:
            task = Task.objects.get(id=pk, user=request.user)
        except Task.DoesNotExist:
            logger.error(f"Task with id {pk} not found for user {request.user}")
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = TaskSerializer(task, data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.debug(f"Task with id {pk} updated successfully")
            return Response(serializer.data)
        
        logger.debug(f"Serializer errors: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        try:
            task = Task.objects.get(id=pk, user=request.user)
        except Task.DoesNotExist:
            logger.error(f"Task with id {pk} not found for user {request.user}")
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
    
        task.delete()
        logger.debug(f"Task with id {pk} deleted successfully")
    
        # Send a success message in the API response
        return Response({"detail": f"Task with id {pk} deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


class ExportTasksToExcelView(APIView):
    authentication_classes = [NodeJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tasks = Task.objects.filter(user=request.user)
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Tasks"
        
        worksheet.append(["ID", "Title", "Description", "Effort (Days)", "Due Date"])

        for task in tasks:
            worksheet.append([task.id, task.title, task.description, task.effort_days, task.due_date])

        for col in range(1, 6):
            max_length = 0
            column = get_column_letter(col)
            for row in worksheet.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Content type for Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename=tasks_{request.user.username}.xlsx'
        
        # Save the workbook to the response
        workbook.save(response)
        return response